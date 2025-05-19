from django.contrib.auth.models import User
from django.utils.text import slugify
from .models import AdminUser
import os
from email.mime.image import MIMEImage
from django.core.mail import EmailMessage
from django.template.loader import render_to_string
from django.conf import settings
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.styles import Font, PatternFill
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from decimal import Decimal


def create_admin_user(full_name, phone_number, role, email):
    """
    Create a new admin user with the provided full_name, phone_number, role, and email.
    """
    if role not in ['admin', 'finance']:
        raise ValueError("Only Admin and Finance roles can be created by admins.")

    # Slugify the full_name to create a unique username (lowercase, hyphenated)
    username = slugify(full_name.lower().replace(" ", "-"))

    # Ensure username is unique
    original_username = username
    counter = 1
    while User.objects.filter(username=username).exists():
        username = f"{original_username}{counter}"
        counter += 1

    # Create User
    user = User.objects.create_user(
        username=username,
        password=phone_number,
        email=email,
        first_name=full_name
    )

    # Create AdminUser profile
    admin_profile = AdminUser.objects.create(
        user=user,
        full_name=full_name,
        email=email,
        role=role,
        phone_number=phone_number
    )

    return admin_profile

def send_membership_email(subject, to_email, context):
    """
    Sends a membership notification email using a shared HTML template.
    """
    html_content = render_to_string('emails/membership_notification.html', context)
    email = EmailMessage(subject, html_content, settings.DEFAULT_FROM_EMAIL, [to_email])
    email.content_subtype = "html"  # Important: Makes it send as HTML
    email.send()

def send_password_reset_email(email, username, reset_link):
    subject = 'Password Reset Request'
    html_content = render_to_string('emails/password_reset.html', {
        'username': username,
        'reset_link': reset_link
    })
    email = EmailMessage(subject, html_content, 'welfare@openedheavenschapel.co.uk', [email])
    email.content_subtype = 'html'
    email.send()

def send_verification_code_email(to_email, code):
    """
    Sends the email verification code using a beautiful HTML template.
    """
    subject = 'Your Email Verification Code'
    context = {
        'username': to_email.split('@')[0],  # You can customize this as needed
        'code': code
    }
    html_content = render_to_string('emails/email_verification_code.html', context)
    email = EmailMessage(subject, html_content, settings.DEFAULT_FROM_EMAIL, [to_email])
    email.content_subtype = "html"
    email.send()

def generate_statement_pdf(member, contributions, from_dt, to_dt):
    from reportlab.lib import colors
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                    Paragraph, Spacer, Image)
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from decimal import Decimal
    import os

    file_name = f"statement_{member.member_id}_{datetime.now().strftime('%Y%m%d%H%M%S')}.pdf"
    folder = os.path.join(settings.MEDIA_ROOT, 'statements')
    os.makedirs(folder, exist_ok=True)
    file_path = os.path.join(folder, file_name)

    doc = SimpleDocTemplate(file_path, pagesize=A4)
    styles = getSampleStyleSheet()

    # 🔥 Custom styles
    title_style = ParagraphStyle(
        'Title',
        parent=styles['Title'],
        fontSize=20,
        spaceAfter=12,
        textColor=colors.HexColor('#6C5CE7')
    )
    normal_bold = ParagraphStyle(
        'NormalBold',
        parent=styles['Normal'],
        fontName='Helvetica-Bold'
    )

    elements = []

    # 1️⃣ LOGO & Title
    logo_path = os.path.join(settings.BASE_DIR, 'static', 'branding', 'heavensconnect_logo.png')
    if os.path.exists(logo_path):
        elements.append(Image(logo_path, width=100, height=100))
    else:
        elements.append(Paragraph("HeavensConnect Welfare App", title_style))

    elements.append(Paragraph(
        f"STATEMENT FOR MEMBER ID: {member.member_id}", title_style))
    elements.append(Paragraph(
        f"Generated on {datetime.now().strftime('%d %B, %Y')}", styles['Normal']))
    elements.append(Spacer(1, 12))

    # 2️⃣ Member Details
    elements.append(Paragraph(f"<b>{member.full_name}</b>", normal_bold))
    elements.append(Paragraph(f"Member ID: {member.member_id}", styles['Normal']))
    if member.address:
        elements.append(Paragraph(member.address, styles['Normal']))
    if member.phone_number:
        elements.append(Paragraph(member.phone_number, styles['Normal']))
    if member.email:
        elements.append(Paragraph(member.email, styles['Normal']))

    elements.append(Spacer(1, 12))

    # 3️⃣ Account Summary
    opening_balance = Decimal('0.00')
    money_in = sum(c.amount for c in contributions if c.amount > 0)
    money_out = Decimal('0.00')  # currently no money out transactions
    closing_balance = money_in - money_out

    summary_data = [
        ['Opening Balance', 'Money Out', 'Money In', 'Closing Balance'],
        [f"£{opening_balance:.2f}", f"£{money_out:.2f}", f"£{money_in:.2f}", f"£{closing_balance:.2f}"]
    ]

    summary_table = Table(summary_data)
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#6C5CE7')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('TOPPADDING', (1, 1), (-1, -1), 6)
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 12))

    # 4️⃣ Statement Period
    elements.append(Paragraph(
        f"Statement period: {from_dt.strftime('%d %B, %Y')} - {to_dt.strftime('%d %B, %Y')}", styles['Normal']))
    elements.append(Spacer(1, 12))

    # 5️⃣ Transactions Table
    data = [['Date', 'Type', 'Description', 'Money Out', 'Money In', 'Balance']]

    running_balance = opening_balance
    for c in contributions:
        date = c.created_at.strftime('%d %b, %Y')
        txn_type = 'Credit'
        desc = f"{c.payment_method.capitalize()} Contribution"
        money_in_txn = c.amount
        money_out_txn = Decimal('0.00')  # Assuming all contributions are credits

        running_balance += money_in_txn

        data.append([
            date,
            txn_type,
            desc,
            f"£{money_out_txn:.2f}",
            f"£{money_in_txn:.2f}",
            f"£{running_balance:.2f}"
        ])

    txn_table = Table(data, colWidths=[1.2 * inch] * 6)
    txn_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#6C5CE7')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('TOPPADDING', (1, 1), (-1, -1), 6)
    ]))
    elements.append(txn_table)

    # 6️⃣ Watermark function
    def add_watermark(canvas_obj, doc_obj):
        canvas_obj.saveState()
        canvas_obj.setFont('Helvetica-Bold', 50)
        canvas_obj.setFillColorRGB(0.9, 0.9, 0.9, alpha=0.2)  # Very light grey
        canvas_obj.drawCentredString(A4[0] / 2, A4[1] / 2, "HeavensConnect")
        canvas_obj.restoreState()

    # Build with watermark on every page
    doc.build(elements, onFirstPage=add_watermark, onLaterPages=add_watermark)

    return file_path

def send_statement_email(member, file_path, from_dt, to_dt):
    subject = f"HeavensConnect Account Statement ({from_dt.strftime('%d %b %Y')} - {to_dt.strftime('%d %b %Y')})"

    context = {
        'full_name': member.full_name,
        'from_date': from_dt.strftime('%d %b %Y'),
        'to_date': to_dt.strftime('%d %b %Y'),
        'support_email': 'welfare@openedheavenschapel.co.uk',
    }

    html_content = render_to_string('emails/account_statement.html', context)

    email = EmailMessage(
        subject,
        html_content,
        'welfare@openedheavenschapel.co.uk',
        [member.email]
    )
    email.content_subtype = "html"
    email.attach_file(file_path)

    # Embed the logo
    logo_path = os.path.join(settings.BASE_DIR, 'static', 'branding', 'heavensconnect_logo.png')
    with open(logo_path, 'rb') as f:
        logo = MIMEImage(f.read())
        logo.add_header('Content-ID', '<logo>')
        email.attach(logo)

    email.send()

def create_notification(user, title, message):
    from .models import Notification  # Avoid circular import
    Notification.objects.create(user=user, title=title, message=message)

    # Fetch FCM token from user profile
    #token = getattr(user, 'fcm_token', None)
    #if token:
    #    send_push_notification(
      #      registration_tokens=[token],
     #       title=title,
      #      body=message
     #   )



# Initialize app only once

def send_finance_report(email, transactions, from_dt, to_dt, export_format='pdf'):
    # 1. Generate the file
    if export_format == 'pdf':
        file_path = generate_finance_report_pdf(transactions, from_dt, to_dt)
    elif export_format == 'excel':
        file_path = generate_finance_report_excel(transactions, from_dt, to_dt)
    else:
        raise ValueError("Unsupported format. Choose 'pdf' or 'excel'.")

    # 2. Compose email
    subject = f"Finance Report – {from_dt.strftime('%d %b %Y')} to {to_dt.strftime('%d %b %Y')}"
    context = {
        'full_name': email.split('@')[0].capitalize(),
        'from_date': from_dt.strftime('%d %b %Y'),
        'to_date': to_dt.strftime('%d %b %Y'),
        'support_email': 'welfare@openedheavenschapel.co.uk',
    }
    html_content = render_to_string('emails/finance_report.html', context)

    message = EmailMessage(
        subject,
        html_content,
        settings.DEFAULT_FROM_EMAIL,
        [email]
    )
    message.content_subtype = "html"
    message.attach_file(file_path)

    # 3. Embed logo
    logo_path = os.path.join(settings.BASE_DIR, 'static', 'branding', 'heavensconnect_logo.png')
    if os.path.exists(logo_path):
        with open(logo_path, 'rb') as f:
            logo = MIMEImage(f.read())
            logo.add_header('Content-ID', '<logo>')
            message.attach(logo)

    message.send()




def generate_finance_report_pdf(transactions, from_dt, to_dt):
    file_name = f"finance_report_{datetime.now().strftime('%Y%m%d%H%M%S')}.pdf"
    folder = os.path.join(settings.MEDIA_ROOT, 'finance_reports')
    os.makedirs(folder, exist_ok=True)
    file_path = os.path.join(folder, file_name)

    doc = SimpleDocTemplate(file_path, pagesize=A4, leftMargin=40, rightMargin=40, topMargin=40, bottomMargin=40)
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        'Title',
        parent=styles['Title'],
        fontSize=20,
        spaceAfter=12,
        textColor=colors.HexColor('#6C5CE7')
    )

    desc_style = ParagraphStyle(
        'desc',
        fontSize=9,
        leading=11,
        spaceAfter=4
    )

    elements = []

    # Title
    elements.append(Paragraph("HEAVENSCONNECT FINANCE REPORT", title_style))
    elements.append(Paragraph(f"Period: {from_dt.strftime('%d %b %Y')} - {to_dt.strftime('%d %b %Y')}", styles['Normal']))
    elements.append(Spacer(1, 12))

    # Summary section
    money_in = sum(txn['amount'] for txn in transactions if txn['type'] == 'income')
    money_out = sum(txn['amount'] for txn in transactions if txn['type'] == 'expense')
    opening_balance = Decimal('0.00')
    closing_balance = opening_balance + Decimal(str(money_in)) - Decimal(str(money_out))

    summary_data = [
        ['Opening Balance', 'Money In', 'Money Out', 'Closing Balance'],
        [f"£{opening_balance:.2f}", f"£{money_in:.2f}", f"£{money_out:.2f}", f"£{closing_balance:.2f}"]
    ]

    summary_table = Table(summary_data)
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#6C5CE7')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (1, 1), (-1, -1), 10),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
        ('TOPPADDING', (0, 0), (-1, 0), 6),
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 18))

    # Transaction Table Header
    data = [['Date', 'Type', 'Amount', 'Description', 'Balance']]
    running_balance = opening_balance

    for txn in transactions:
        if isinstance(txn['date'], str):
            txn['date'] = datetime.fromisoformat(txn['date'].replace('Z', '+00:00'))

        date = txn['date'].strftime('%d %b %Y')
        type_label = txn['type'].capitalize()
        amount = txn['amount']
        description_text = txn.get('description', 'N/A')
        description = Paragraph(description_text, desc_style)
        running_balance += amount if txn['type'] == 'income' else -amount
        balance = f"£{running_balance:.2f}"

        data.append([date, type_label, f"£{amount:.2f}", description, balance])

    # Adjusted column widths
    col_widths = [1.2 * inch, 0.9 * inch, 1.1 * inch, 3.4 * inch, 1.2 * inch]

    txn_table = Table(data, colWidths=col_widths, repeatRows=1)
    txn_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#6C5CE7')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (2, -1), 'CENTER'),  # Center for Date, Type, Amount
        ('ALIGN', (3, 1), (3, -1), 'LEFT'),    # Description left-aligned
        ('ALIGN', (4, 1), (4, -1), 'RIGHT'),   # Balance right-aligned
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 0.4, colors.grey),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
        ('TOPPADDING', (0, 0), (-1, 0), 6),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    elements.append(txn_table)

    # Watermark
    def add_watermark(canvas_obj, doc_obj):
        canvas_obj.saveState()
        canvas_obj.setFont('Helvetica-Bold', 50)
        canvas_obj.setFillColorRGB(0.9, 0.9, 0.9, alpha=0.1)
        canvas_obj.drawCentredString(A4[0] / 2, A4[1] / 2, "HeavensConnect")
        canvas_obj.restoreState()

    doc.build(elements, onFirstPage=add_watermark, onLaterPages=add_watermark)
    return file_path

def generate_finance_report_excel(transactions, from_dt, to_dt):
    file_name = f"finance_report_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
    folder = os.path.join(settings.MEDIA_ROOT, 'finance_reports')
    os.makedirs(folder, exist_ok=True)
    file_path = os.path.join(folder, file_name)

    wb = Workbook()
    ws = wb.active
    ws.title = "Finance Report"

    headers = ['Date', 'Type', 'Amount', 'Description', 'Balance']
    header_font = Font(bold=True, color="FFFFFF")
    fill = PatternFill(start_color="6C5CE7", end_color="6C5CE7", fill_type="solid")

    ws.append(headers)
    for col in range(1, len(headers) + 1):
        ws.cell(row=1, column=col).font = header_font
        ws.cell(row=1, column=col).fill = fill

    running_balance = Decimal('0.00')

    for txn in transactions:
        if isinstance(txn['date'], str):
            txn['date'] = datetime.fromisoformat(txn['date'].replace('Z', '+00:00'))

        date = txn['date'].strftime('%d %b %Y')
        type_label = txn['type'].capitalize()
        amount = Decimal(str(txn['amount']))  # Ensure Decimal
        description = txn.get('description', 'N/A').capitalize()

        if txn['type'] == 'income':
            running_balance += amount
        else:
            running_balance -= amount

        ws.append([
            date,
            type_label,
            f"£{amount:.2f}",
            description,
            f"£{running_balance:.2f}"
        ])

    # Auto-size columns
    for column_cells in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = max_length + 2

    wb.save(file_path)
    return file_path
